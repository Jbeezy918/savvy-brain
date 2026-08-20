import re
from pathlib import Path
from openpyxl import load_workbook

master = Path.home() / "Desktop/Savvy Master Library/Savvy_Master.xlsx"
wb = load_workbook(master)
keys = wb["API Keys & Tokens"]

if "ENV Export" in wb.sheetnames:
    del wb["ENV Export"]

env = wb.create_sheet("ENV Export")
env.append([
    "Include", "Project", "Environment", "Variable Name",
    "Full Value", "Provider", "Credential Type",
    "Format Status", "Source File", "Notes"
])

headers = {
    str(cell.value).strip(): index
    for index, cell in enumerate(keys[1], start=1)
}

used = set()

for row in range(2, keys.max_row + 1):
    provider = str(keys.cell(row, headers["Provider"]).value or "").strip()
    credential_type = str(
        keys.cell(row, headers["Credential Type"]).value or ""
    ).strip()
    value = str(keys.cell(row, headers["Full Value"]).value or "").strip()
    source = str(keys.cell(row, headers["Source File"]).value or "").strip()
    variable = str(
        keys.cell(row, headers["Environment Variable"]).value or ""
    ).strip()

    variable = re.sub(r"[^A-Za-z0-9_]+", "_", variable.upper()).strip("_")
    variable = variable or "UNKNOWN_SECRET"

    original = variable
    number = 2
    while variable in used:
        variable = f"{original}_{number}"
        number += 1
    used.add(variable)

    status = "PLAUSIBLE — live test not performed"
    if not value:
        status = "INVALID — empty value"
    elif any(character.isspace() for character in value):
        status = "INVALID — contains whitespace"

    env.append([
        "YES" if not status.startswith("INVALID") else "NO",
        "General",
        "local",
        variable,
        value,
        provider,
        credential_type,
        status,
        source,
        "Set Include to YES or NO, then save workbook."
    ])

env.freeze_panes = "A2"
env.auto_filter.ref = env.dimensions
wb.save(master)

print("ENV Export tab created.")
print(master)
