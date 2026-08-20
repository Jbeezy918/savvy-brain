from pathlib import Path
from openpyxl import load_workbook

master = Path.home() / "Desktop/Savvy Master Library/Savvy_Master.xlsx"
output_dir = Path.home() / "Desktop/Savvy Master Library/ENV Exports"
output_file = output_dir / "Savvy_Master.env"

wb = load_workbook(master, data_only=False)
sheet = wb["ENV Export"]

headers = {
    str(cell.value).strip(): index
    for index, cell in enumerate(sheet[1], start=1)
}

lines = [
    "# Savvy Master ENV Export",
    "# Keep private. Never commit this file to Git.",
    "",
]

count = 0

for row in range(2, sheet.max_row + 1):
    include = str(
        sheet.cell(row, headers["Include"]).value or ""
    ).strip().upper()

    if include not in {"YES", "Y", "TRUE", "1"}:
        continue

    variable = str(
        sheet.cell(row, headers["Variable Name"]).value or ""
    ).strip()
    value = str(
        sheet.cell(row, headers["Full Value"]).value or ""
    ).strip()
    status = str(
        sheet.cell(row, headers["Format Status"]).value or ""
    ).strip()

    if not variable or not value or status.startswith("INVALID"):
        continue

    escaped = value.replace("\\", "\\\\").replace('"', '\\"')
    lines.append(f'{variable}="{escaped}"')
    count += 1

output_dir.mkdir(parents=True, exist_ok=True)
output_file.write_text("\n".join(lines) + "\n", encoding="utf-8")
output_file.chmod(0o600)

print(f"Exported {count} credential(s).")
print(output_file)
